"""Genera un archivo Excel de ejemplo con órdenes de producción."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def crear_excel_ejemplo():
    """Crea un archivo Excel con órdenes de ejemplo."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ordenes"

    ws.merge_cells('A1:G1')
    ws['A1'] = "Órdenes de Producción - Área de Prensas (con pisos)"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')

    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = ["id", "dim_x", "dim_y", "cantidad", "tiempo_unitario", "presion_psi", "prioridad"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Órdenes diseñadas para demostrar emparejamiento
    ordenes = [
        # Par 1: Tiempos similares (1.0 y 1.2 h), presiones cercanas (100 y 110 PSI)
        ("ORD-001", 1000, 500, 100, 1.0, 100, 1),
        ("ORD-002", 1100, 550, 80, 1.2, 110, 1),

        # Par 2: Tiempos similares (1.5 y 1.8 h), presiones cercanas (120 y 130 PSI)
        ("ORD-003", 1200, 580, 100, 1.5, 120, 2),
        ("ORD-004", 1150, 570, 90, 1.8, 130, 2),

        # Par 3: Tiempos similares (2.0 y 2.2 h), presiones cercanas (105 y 115 PSI)
        ("ORD-005", 1400, 550, 60, 2.0, 105, 3),
        ("ORD-006", 1350, 520, 55, 2.2, 115, 3),

        # Ordenes sin emparejar
        ("ORD-007", 1700, 700, 40, 1.5, 200, 1),  # Presión alta
        ("ORD-008", 1900, 1900, 20, 2.0, 150, 2),  # Prensa P7 específica

        # Más pares
        ("ORD-009", 1000, 600, 120, 0.8, 100, 1),
        ("ORD-010", 1100, 590, 100, 1.0, 105, 1),

        ("ORD-011", 1500, 500, 50, 1.6, 125, 2),
        ("ORD-012", 1450, 520, 45, 2.0, 130, 2),
    ]

    for row_idx, orden in enumerate(ordenes, 4):
        for col_idx, valor in enumerate(orden, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=valor)
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border

    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 18
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10

    ws_notes = wb.create_sheet(title="Notas")
    ws_notes['A1'] = "Notas sobre el formato de órdenes:"
    ws_notes['A1'].font = Font(bold=True)

    notas = [
        "",
        "COLUMNAS REQUERIDAS:",
        "- id: Identificador único de la orden",
        "- dim_x: Dimensión X de la pieza en milímetros",
        "- dim_y: Dimensión Y de la pieza en milímetros",
        "- cantidad: Cantidad a producir",
        "- tiempo_unitario: Tiempo de prensado por pieza en horas",
        "",
        "COLUMNAS OPCIONALES:",
        "- presion_psi: Presión de prensado en PSI (para emparejamiento)",
        "- prioridad: Nivel de prioridad (1=alta, 5=baja). Default: 3",
        "",
        "PRENSAS CON PISOS:",
        "- P1: 1200x600mm, 50 EUR/h (2 pisos)",
        "- P2: 1200x600mm, 50 EUR/h (2 pisos)",
        "- P5: 1500x600mm, 60 EUR/h (2 pisos)",
        "- P6: 1800x800mm, 75 EUR/h (2 pisos)",
        "- P7: 2000x2000mm, 100 EUR/h (2 pisos)",
        "",
        "RESTRICCIONES DE EMPAREJAMIENTO:",
        "- Diferencia máxima de tiempo: 50% del menor tiempo",
        "- Diferencia máxima de presión: 20 PSI",
        "",
        "EJEMPLO DE EMPAREJAMIENTO VÁLIDO:",
        "- ORD-001: 1.0 h, 100 PSI con ORD-002: 1.2 h, 110 PSI",
        "  (Diferencia tiempo: 20%, presión: 10 PSI) ✓",
    ]

    for idx, nota in enumerate(notas, 2):
        ws_notes.cell(row=idx, column=1, value=nota)

    ws_notes.column_dimensions['A'].width = 70

    wb.save("datos/ejemplo_ordenes.xlsx")
    print("Archivo creado: datos/ejemplo_ordenes.xlsx")


if __name__ == "__main__":
    import os
    os.makedirs("datos", exist_ok=True)
    crear_excel_ejemplo()
